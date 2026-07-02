# -*- coding: utf-8 -*-
"""
배급 시간표 엑셀(그린랜드2 마이그레이션_시간표*.xlsx)에서 오늘 날짜 시트의
- 총 상영회차 / 총 좌석수
- 시간대별(오전 ~12:00 / 오후 12:01~17:00 / 저녁 17:01~) 회차 수
를 뽑아 schedule.json 저장 → 대시보드 '편성 반영 예상 스코어'에 사용.
(시간표는 수동 스냅샷; 0701~0705 등 여러 날짜 시트 포함)
"""
import os
import re
import glob
import json
import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
OUT_JSON = os.path.join(BASE, "schedule.json")
SEARCH_DIRS = [os.path.join(os.path.expanduser("~"), "Downloads"),
               os.path.join(os.path.expanduser("~"), "OneDrive - 키노라이츠", "바탕 화면"),
               BASE]


def _find_file():
    cands = []
    for d in SEARCH_DIRS:
        cands += glob.glob(os.path.join(d, "*마이그레이션*시간표*.xls*"))
        cands += glob.glob(os.path.join(d, "*시간표*.xlsx"))
    return max(cands, key=os.path.getmtime) if cands else None


def _num(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def parse(path, date_str=None):
    import openpyxl
    if date_str is None:
        date_str = datetime.date.today().strftime("%Y-%m-%d")
    mmdd = date_str[5:7] + date_str[8:10]  # "0701"
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = mmdd if mmdd in wb.sheetnames else next(
        (s for s in wb.sheetnames if re.fullmatch(r"\d{4}", s)), wb.sheetnames[0])
    ws = wb[sheet]
    rows = [list(r) for r in ws.iter_rows(values_only=True)]

    bands = {"오전": 0, "오후": 0, "저녁": 0}
    total_seats = 0
    total_screens = None
    # 체인별 편성(극장/상영관/상영회차/좌석수) — '합계' 행에서
    CHAINMAP = {"롯데": "롯데시네마", "CGV": "CGV", "메가": "메가박스"}
    chains = {}
    cur_chain = None
    for r in rows:
        cells = [("" if c is None else c) for c in r]
        c0 = str(cells[0]).strip() if cells else ""
        c1 = str(cells[1]).strip() if len(cells) > 1 else ""
        if c0 in CHAINMAP:
            cur_chain = CHAINMAP[c0]
        if c1 == "합계" and cur_chain and len(cells) > 6:
            chains[cur_chain] = {"극장": _num(cells[2]), "회차": _num(cells[3]),
                                 "상영관": _num(cells[5]), "좌석": _num(cells[6])}
            cur_chain = None

    for r in rows:
        cells = [("" if c is None else c) for c in r]
        first = str(cells[0]).strip() if cells else ""
        # 시간대 비율표: 첫 셀이 오전/오후/저녁, 정수(>1)들의 합 = 그 시간대 회차수
        if first in bands:
            # 행에 계열사별 회차 + 총합 컬럼이 함께 있음 → 최댓값(=그 시간대 총 회차)만 사용
            ints = [n for c in cells[1:]
                    if (n := _num(c)) is not None and n > 1 and float(c) == int(float(c))]
            if ints:
                bands[first] = max(ints)
        # 총계: '계' 행(좌석 큰 값)에서 좌석/상영관
        if first == "계":
            seats_here = max((_num(c) or 0) for c in cells)
            if seats_here > 10000:
                total_seats = seats_here
                total_screens = _num(cells[5]) if len(cells) > 5 else None
    total_shows = sum(bands.values())

    hourly = _parse_hourly(rows)
    regions = _parse_regions(rows)
    return {"date": date_str, "sheet": sheet, "total_shows": total_shows,
            "total_seats": total_seats, "total_screens": total_screens,
            "bands": bands, "chains": chains, "hourly": hourly, "regions": regions}


# 시간대 세부표: 밴드 라벨 col3, 계열사 col2, 회차 카운트 col7부터.
# 밴드 내 상대 슬롯 → 절대 시각 매핑(오전은 단일값이라 11시로 집약).
_BAND_HOURS = {
    "오전": [11],
    "오후": [12, 13, 14, 15, 16],
    "저녁": [17, 18, 19, 20, 21, 22, 23],
}


def _parse_hourly(rows):
    """계열사×밴드 시간분포표 → {절대시각(str): 회차수} (전 계열사 합)."""
    hourly = {}
    for r in rows:
        cells = [("" if c is None else c) for c in r]
        if len(cells) < 8:
            continue
        c3 = str(cells[3]).strip()
        band = next((b for b in _BAND_HOURS if c3.startswith(b)), None)
        if not band:
            continue
        hrs = _BAND_HOURS[band]
        if band == "오전":
            counts = [sum((_num(c) or 0) for c in cells[7:18])]
        else:
            counts = [(_num(c) or 0) for c in cells[7:7 + len(hrs)]]
        for h, n in zip(hrs, counts):
            if n:
                hourly[h] = hourly.get(h, 0) + n
    return {str(k): v for k, v in sorted(hourly.items())}


def _parse_regions(rows):
    """우측 지역별 블록 → [[지역, 극장수, 회차, 상영관, 좌석], ...]."""
    out = []
    for r in rows:
        cells = [("" if c is None else c) for c in r]
        if len(cells) < 21:
            continue
        name = str(cells[10]).strip()
        seats = _num(cells[20])
        # 지역명은 문자열(숫자 아님) + 좌석 큰 값 → 시간분포표 오염행 배제
        if (not name or name in ("지역별", "계") or _num(name) is not None
                or seats is None or seats <= 1000):
            continue
        out.append([name, _num(cells[12]), _num(cells[14]), _num(cells[18]), seats])
    return out


HIST_JSON = os.path.join(BASE, "schedule_history.json")
CAP_LOG = os.path.join(BASE, "schedule_capacity_log.json")  # 좌석 개방 곡선(수집일×대상일)


def main():
    import openpyxl
    f = _find_file()
    if not f:
        print("시간표 파일 못 찾음")
        return 1
    data = parse(f)
    with open(OUT_JSON, "w", encoding="utf-8") as fp:
        json.dump(data, fp, ensure_ascii=False)
    # 날짜별 이력(모든 날짜 시트) 저장 → 대시보드 날짜 선택용
    hist = {}
    if os.path.exists(HIST_JSON):
        try:
            hist = json.load(open(HIST_JSON, encoding="utf-8"))
        except Exception:
            hist = {}
    # 좌석 개방 곡선(수집일 asof × 대상일) — 주말 좌석이 며칠에 걸쳐 열리는 추이
    cap = {}
    if os.path.exists(CAP_LOG):
        try:
            cap = json.load(open(CAP_LOG, encoding="utf-8"))
        except Exception:
            cap = {}
    asof = datetime.date.today().strftime("%Y-%m-%d")

    year = datetime.date.today().strftime("%Y")
    for s in openpyxl.load_workbook(f, read_only=True).sheetnames:
        if re.fullmatch(r"\d{4}", s):
            ds = f"{year}-{s[:2]}-{s[2:]}"
            try:
                d = parse(f, ds)
                if d.get("total_seats"):
                    hist[ds] = {"chains": d["chains"], "total_seats": d["total_seats"],
                                "total_shows": d["total_shows"], "total_screens": d.get("total_screens"),
                                "hourly": d.get("hourly", {}), "regions": d.get("regions", [])}
                    cap.setdefault(ds, {})[asof] = {
                        "seats": d["total_seats"], "shows": d["total_shows"],
                        "screens": d.get("total_screens")}
            except Exception:
                pass
    json.dump(hist, open(HIST_JSON, "w", encoding="utf-8"), ensure_ascii=False)
    json.dump(cap, open(CAP_LOG, "w", encoding="utf-8"), ensure_ascii=False)
    print("시간표 파싱:", data.get("date"), "| 이력 날짜:", sorted(hist), "| 좌석로그 asof:", asof)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
